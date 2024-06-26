
from openpyxl.styles import Font, PatternFill, numbers, DEFAULT_FONT, Color, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell

from models.excel_base import ExcelBase
from report_tools.line_item_position_config import ITEM_POSITION


class ExcelReport(ExcelBase):
    def __init__(self, file_name: str = None):
        super().__init__(file_name)
        self.sheet_names = [
            "Transport",
            "Materials",
        ]
        self.colors = {
            "transport": "00FF9900",
            "materials": "0099CC00",
            "header": "00E4DFEC",
            "default": "0099CCFF",
            "calculate": "00F7F7F7",
            "tariff_fill": "00D8E4BC",
        }
        self.fonts = {
            "default": Font(name="Arial", size=8, bold=False, italic=False),
            "default_bold": Font(name="Arial", size=8, bold=True, italic=False),
            "green_bold": Font(name="Arial", size=8, bold=True, color="006600"),
            "grey": Font(name="Arial", size=8, bold=False, color=Color("808080")),
            "blue": Font(name="Arial", size=8, bold=False, color="1F497D"),
            "result_bold": Font(name="Arial", size=8, bold=True, color=Color("990000")),
        }
        self.fills = {
            "calculate": PatternFill("solid", fgColor=self.colors["calculate"]),
            "header": PatternFill("solid", fgColor=self.colors["header"]),
            "tariff": PatternFill("solid", fgColor=self.colors["tariff_fill"]),
        }
        self.alignments= {
            "default": Alignment(horizontal='right', vertical='bottom'),
            "left": Alignment(horizontal='left', vertical='bottom'),
            "flag_char": Alignment(horizontal='center', vertical='center'),
        }

        self.number_format = "#,##0.00"
        self.counter_format = "#,##0"

    #
    def __enter__(self):
        super().__enter__()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            # self.save_file()
            super().__exit__(exc_type, exc_val, exc_tb)


    def get_sheet(self, sheet_name: str, index: int = None):
        """удаляет лист в книге и создает новый"""
        if sheet_name in (x.title for x in self.workbook.worksheets):
            self.workbook.remove(self.workbook[sheet_name])
        if index is None:
            self.worksheet = self.workbook.create_sheet(
                sheet_name, index=len(self.workbook.worksheets)
            )
        else:
            self.worksheet = self.workbook.create_sheet(sheet_name, index=index)
        self.worksheet.font = self.fonts["default"]
        if sheet_name in self.colors.keys():
            self.worksheet.sheet_properties.tabColor = self.colors[sheet_name]
        else:
            self.worksheet.sheet_properties.tabColor = self.colors["default"]
        return self.worksheet

    def activate_sheet(self, sheet_name: str, index: int = None):
        """Активирует лист в рабочей книге по его имени."""
        if sheet_name in (ws.title for ws in self.workbook.worksheets):
            self.worksheet = self.workbook[sheet_name]
            self.workbook.active = self.worksheet
        return self.worksheet



    def create_sheets(self):
        """удаляет все листы в книге и создает новые"""
        if self.workbook:
            for sheet in self.workbook.worksheets:
                self.workbook.remove(sheet)
            # DEFAULT_FONT = self.fonts["default"]
            for i, name in enumerate(self.sheet_names):
                self.sheet = self.workbook.create_sheet(name)
                self.sheet.font = self.fonts["default"]
                self.sheet.sheet_properties.tabColor = self.colors[name]

    def delete_sheets(self, sheets_to_delete: list[str]):
        """Удаляет листы из рабочей книги"""
        for sheet_name in sheets_to_delete:
            if sheet_name in (x.title for x in self.workbook.worksheets):
                self.workbook.remove(self.workbook[sheet_name])

    def set_sheet_number(self, sheet_name: str, number: int):
        """Устанавливает номер листа"""
        if sheet_name in (x.title for x in self.workbook.worksheets):
            sheet_index = self.workbook.worksheets.index(sheet_name)
            self.workbook.worksheets[sheet_index] = self.workbook.worksheets.pop(sheet_index)
            self.workbook.worksheets.insert(number - 1, self.worksheet)

    def write_header(self, sheet_name: str, header: list, row: int = 1):
        """записывает заголовок в лист"""
        # self.create_sheets()
        self.worksheet = self.workbook[sheet_name]
        # self.worksheet.append(header)

        for col in range(1, len(header) + 1):
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = header[col - 1]
            cell.font = self.fonts["default"]
            cell.fill = self.fills["header"]
            cell.alignment = Alignment(wrap_text=True, vertical="bottom")

    def write_row(self, sheet_name: str, values: list, row_index: int = 2):
        """записывает строку в лист"""
        worksheet = self.workbook[sheet_name]
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index).value = value
            # worksheet.cell(row=row_index, column=column_index).font = self.font

    def write_format(self, sheet_name: str, row_index: int, len_row):
        """записывает строку в лист"""

        self.worksheet = self.workbook[sheet_name]
        for column_index in range(14, 18):
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["default"]
            cell.number_format = self.number_format

        self.worksheet.cell(row=row_index, column=1).font = self.fonts["default_bold"]

        price_cells = [2, 10, 12]
        for column_index in price_cells:
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["green_bold"]
            cell.number_format = self.number_format

        for col in range(3, 9):
            cell = self.worksheet.cell(row=row_index, column=col)
            cell.font = self.fonts["grey"]

        index_cells = (9, 11)
        for column_index in index_cells:
            cell = self.worksheet.cell(row=row_index, column=column_index)
            cell.font = self.fonts["blue"]


    def set_cell_format(
        self, row: int, column_name: str,
        font_name: str = None, number_format: str = None,
        alignment: Alignment = None, fill: PatternFill = None
    ):
        """Устанавливает цвет и формат ячейки"""
        column_number = ITEM_POSITION[column_name].column_number
        cell = self.worksheet.cell(row=row, column=column_number)
        
        if font_name:
            cell.font = self.fonts[font_name]
        if number_format:
            cell.number_format = number_format
        if alignment:
            cell.alignment = alignment
        else:
            cell.alignment = self.alignments["default"]
        if fill:
            cell.fill = fill

    def _set_column_width(self, column_name: str, width: int):
        """Устанавливает ширину столбца"""
        column_letter = ITEM_POSITION[column_name].column_letter
        self.worksheet.column_dimensions[column_letter].width = width

    



    def write_material_format(self, sheet_name: str, row_index: int, history_len: int):
        """"""
        self.set_cell_format(row_index, "row_count", "grey", "# ##0")
        self.set_cell_format(row_index, "code", "default_bold")
        self.set_cell_format(row_index, "name", "default", alignment=self.alignments["left"])
        self.set_cell_format(row_index, "base_price", "default", self.number_format)
        #  history
        start_col = ITEM_POSITION["price_history_range"].column_number
        for col in range(start_col, start_col + history_len):
            cell = self.worksheet.cell(row=row_index, column=col)
            cell.font = self.fonts["grey"]
            cell.number_format = self.number_format

        # 
        self.set_cell_format(row_index, "last_period_delivery", "default", alignment=self.alignments["flag_char"])
        self.set_cell_format(row_index, "check_need", "default")
        self.set_cell_format(row_index, "supplier_price", "green_bold", self.number_format)
        self.set_cell_format(row_index, "is_delivery_included", "default_bold", alignment=self.alignments["flag_char"])
        # 
        self.set_cell_format(row_index, "transport_code", "grey")
        self.set_cell_format(row_index, "transport_base_price", "grey", self.number_format)
        self.set_cell_format(row_index, "transport_numeric_ratio", "grey", self.number_format)
        self.set_cell_format(row_index, "transport_actual_price", "grey", self.number_format)
        #
        self.set_cell_format(row_index, "gross_weight", "default", self.number_format)
        self.set_cell_format(row_index, "unit_measure", "default")
        self.set_cell_format(row_index, "current_selling_price", "blue", self.number_format)
        self.set_cell_format(row_index, "empty_1", "default")
        self.set_cell_format(row_index, "transport_price", "default_bold", self.number_format)
        # 
        self.set_cell_format(row_index, "result_price", "result_bold", self.number_format)
        self.set_cell_format(row_index, "previous_index", "default", self.number_format)
        self.set_cell_format(row_index, "result_index", "default", self.number_format)
        # 
        self.set_cell_format(row_index, "index_change_absolute", "default", self.number_format)
        self.set_cell_format(row_index, "index_change_in_percentage", "default", "0.00%")
        # 
        self.set_cell_format(row_index, "absolute_price_change", "default", self.number_format)
        self.set_cell_format(row_index, "percentage_price_change", "default", "0.00%")
        # 
        slime_cells = ["empty_1", "empty_2", "row_count"]
        for column_name in slime_cells:        
            self._set_column_width(column_name, 3)
        # 
        self._set_column_width("unit_measure", 5)
        self._set_column_width("last_period_delivery", 7)
        self._set_column_width("is_delivery_included", 7)
        # 
        calculate_cells = [
            "transport_price", "result_price", "previous_index", "result_index", "index_change_absolute", 
            "index_change_in_percentage", "absolute_price_change", "percentage_price_change","empty_2",]
        for column_name in calculate_cells:        
            self.set_cell_format(row_index, column_name, fill=self.fills["calculate"])



    def set_monitoring_price_header_format(
        self, sheet_name: str, row: int = 1, header: list = None
    ):
        """Устанавливает формат строки заголовка отчета о ценах мониторинга."""
        text_length = 4
        sheet = self.workbook[sheet_name]
        # Установите формат для первых 3 столбцов
        for col in range(1, text_length + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = self.fonts["default"]
            cell.fill = self.fills["header"]
            cell.alignment = Alignment(wrap_text=True)
        # Установите формат для остальных столбцов
        for col in range(text_length+1, len(header) + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = self.fonts["default_bold"]
            cell.fill = self.fills["header"]
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

    def set_regular_row_monitoring_price_format(
        self,
        sheet_name: str,
        row: int,
        price_delivery_flags: tuple[bool,...] = None,
    ):
        """Устанавливает формат для строки мониторинга цен."""
        counter_col = 1
        code_col = 2
        delivery_col = 3
        title_col = 4
        price_start_col = 5
        sheet = self.workbook[sheet_name]
        counter_cell = sheet.cell(row=row, column=counter_col)
        code_cell = sheet.cell(row=row, column=code_col)
        delivery_cell = sheet.cell(row=row, column=delivery_col)
        title_cell = sheet.cell(row=row, column=title_col)
        # counter format
        counter_cell.font = self.fonts["grey"]
        counter_cell.number_format = self.counter_format
        counter_cell.alignment = Alignment(horizontal="center")
        # code format
        if True in price_delivery_flags:
            code_cell.font = self.fonts["result_bold"]
            code_cell.alignment = Alignment(horizontal="left")
            # delivery format
            delivery_cell.font = self.fonts["result_bold"]
            delivery_cell.number_format = self.counter_format
            delivery_cell.alignment = Alignment(horizontal="center")
        else:
            code_cell.font = self.fonts["default_bold"]
            delivery_cell.font = self.fonts["grey"]
        # title format
        title_cell.font = self.fonts["default"]
        # price format
        for col, flag in enumerate(price_delivery_flags, start=price_start_col):
            price_cell = sheet.cell(row=row, column=col)
            if flag:
                price_cell.font = self.fonts["result_bold"]
            else:
                price_cell.font = self.fonts["default"]
            price_cell.number_format = self.number_format

    def set_column_widths(self, sheet_name: str, width: int, columns: tuple[int, ...] = None) -> None:
        """Устанавливает ширину столбцов в заданном листе."""
        sheet = self.workbook[sheet_name]
        for column_index in columns:
            sheet.column_dimensions[get_column_letter(column_index)].width = width


    def set_tariffs_header_format(
        self, sheet_name: str, row: int = 1, header_len: int = 1):
        """Устанавливает формат строки заголовка отчета о ценах на тарифы."""
        sheet = self.workbook[sheet_name]
        for col in range(1, header_len + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = self.fonts["default"]
            cell.fill = self.fills["header"]
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

    def set_regular_row_tariff_format(self, sheet_name: str, row: int = 1):
        """Устанавливает формат для строки цен на тарифы."""
        sheet = self.workbook[sheet_name]
        # text format
        text_columns = [1, 4, 5]
        for col in text_columns:
            cell = sheet.cell(row=row, column=col)
            cell.font = self.fonts["default"]
            cell.alignment = Alignment(wrap_text=False, horizontal="left")
        #
        num_cell = sheet.cell(row=row, column=2)
        num_cell.font = self.fonts["default"]
        num_cell.alignment = Alignment(wrap_text=True, horizontal="center")
        num_cell.number_format = self.counter_format
        #
        code_cell = sheet.cell(row=row, column=3)
        code_cell.font = self.fonts["default_bold"]
        code_cell.alignment = Alignment(wrap_text=True, horizontal="right")
        #
        price_cell = sheet.cell(row=row, column=6)
        price_cell.font = self.fonts["result_bold"]
        price_cell.alignment = Alignment(wrap_text=True, horizontal="right")
        price_cell.number_format = self.number_format

    def format_material_prices_per_tariffs(self, format_cell: Cell):
        """Форматирует ячейку с ценой материала измененной по тарифу."""
        format_cell.fill = self.fills["tariff"]
        format_cell.alignment = Alignment(horizontal="right")