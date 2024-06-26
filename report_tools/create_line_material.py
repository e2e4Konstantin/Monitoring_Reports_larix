from typing import Any
from icecream import ic
from openpyxl.utils import get_column_letter
from models import MonitoringMaterial, MonitoringPrice
import itertools

from report_tools.line_item_position_config import ITEM_POSITION


def _create_price_history_range(
    monitoring_prices: list[MonitoringPrice] = None, max_length: int = 0
) -> list[float | None]:
    """Создайте список цен, историю последних max_len периодов. 
    Если длина меньше max_length, заполнит None."""
    empty_range = [None] * max_length
    if not monitoring_prices:
        return empty_range
    prices = [price.price for price in monitoring_prices]
    if len(prices) >= max_length:
        return prices[-max_length:]
    return empty_range[:-len(prices)] + prices

def _create_delivery_included_history(
    monitoring_prices: list[MonitoringPrice] = None, max_length: int = 0
) -> list[float | None]:
    """Создайте список признаков включения транспортных расходов в цену мониторинга, 
    историю последних max_len периодов. 
    Если длина меньше max_length, заполнит None."""
    empty_range = [None] * max_length
    if not monitoring_prices:
        return empty_range
    delivery = [price.delivery for price in monitoring_prices]
    if len(delivery) >= max_length:
        return delivery[-max_length:]
    return empty_range[:-len(delivery)] + delivery



def _set_new_value(item_name: str, field_to_update: str, value: Any) -> None:
    """Устанавливает значение  в словаре ITEM_POSITION для заданного элемента."""
    if hasattr(ITEM_POSITION[item_name], field_to_update):
        ITEM_POSITION[item_name] = ITEM_POSITION[item_name]._replace(**{field_to_update: value})


def _set_items_position() -> None:
    """ 
    Устанавливает значение позиций элементов в словаре ITEM_POSITION.
    Назначает букву для столбца
    """
    price_range = ITEM_POSITION["price_history_range"].value
    history_end_column = ITEM_POSITION["price_history_range"].column_number + len(price_range)-1
    # 
    _set_new_value("last_period_delivery", "column_number", history_end_column + 1) 
    # 
    _set_new_value("check_need", "column_number", history_end_column + 2)
    _set_new_value("supplier_price", "column_number", history_end_column + 3)
    _set_new_value("is_delivery_included", "column_number", history_end_column + 4)
    _set_new_value("transport_code", "column_number", history_end_column+5)
    _set_new_value("transport_base_price", "column_number", history_end_column+6)
    _set_new_value("transport_numeric_ratio", "column_number", history_end_column+7)
    _set_new_value("transport_actual_price", "column_number", history_end_column+8)
    _set_new_value("gross_weight", "column_number", history_end_column+9)
    _set_new_value("unit_measure", "column_number", history_end_column+10)
    _set_new_value("current_selling_price", "column_number", history_end_column+11)
    _set_new_value("empty_1", "column_number", history_end_column+12)
    # 
    _set_new_value("transport_price", "column_number", history_end_column+13)
    _set_new_value("result_price", "column_number", history_end_column+14)
    _set_new_value("previous_index", "column_number", history_end_column+15)
    _set_new_value("result_index", "column_number", history_end_column+16)
    _set_new_value("index_change_absolute", "column_number", history_end_column+17)
    _set_new_value("index_change_in_percentage", "column_number", history_end_column+18)
    # 
    _set_new_value("empty_2", "column_number", history_end_column+19)
    _set_new_value("absolute_price_change", "column_number", history_end_column+20)
    _set_new_value("percentage_price_change", "column_number", history_end_column+21)
    # добавляем букву для номера колонки
    for key in ITEM_POSITION.keys():
        _set_new_value(key, "column_letter", get_column_letter(ITEM_POSITION[key].column_number))




def create_line_material(material: MonitoringMaterial, row_number: int, max_history_len: int):
    """ Создать строку с данными о материале."""
    index_data = material.index_period_material_data
    _set_new_value("row_count", "value", row_number)
    _set_new_value("code", "value", material.code)
    _set_new_value("name", "value", material.description)
    _set_new_value("base_price", "value", index_data.base_price)
    # 
    # delivery_range = _create_delivery_included_history(
    #     material.monitoring_price_history, max_history_len
    # )
    # 
    price_range = _create_price_history_range(
        material.monitoring_price_history, max_history_len
    )
    _set_new_value("price_history_range", "value",  price_range)
    # !!! назначить столбцы
    _set_items_position()
     
    value = "+" if material.monitoring_price_history[-1].delivery else ""
    _set_new_value("last_period_delivery", "value", value)
    # 
    # value = 1 if material.is_delivery_included != material.monitoring_price_history[-1].delivery else ""
    # _set_new_value("check_need", "value", value)
    # 
    _set_new_value("supplier_price", "value", material.supplier_price)
    _set_new_value("is_delivery_included", "value", "+" if material.is_delivery_included else "")
    _set_new_value("transport_code", "value", index_data.transport_code)
    _set_new_value("transport_base_price", "value", index_data.transport_base_price)
    _set_new_value("transport_numeric_ratio", "value", index_data.transport_inflation_rate)
    _set_new_value("transport_actual_price", "value", index_data.transport_current_price)
    _set_new_value("gross_weight", "value", index_data.gross_weight)
    _set_new_value("unit_measure", "value", material.unit_measure)
    _set_new_value("current_selling_price", "value", index_data.current_price)
    _set_new_value("empty_1", "value", "")
    # формул
    _set_new_value("transport_price", "value", "")
    _set_new_value("result_price", "value", "")
    _set_new_value("previous_index", "value", "")
    _set_new_value("result_index", "value", "")
    _set_new_value("absolute_price_change", "value", "")
    _set_new_value("percentage_price_change", "value", "")
    # формулы 
    is_delivery_included_column = ITEM_POSITION["is_delivery_included"].column_letter
    last_period_delivery_column = ITEM_POSITION["last_period_delivery"].column_letter
    check_need_formula = (
        f'=IF({last_period_delivery_column}{row_number}<>{is_delivery_included_column}{row_number}, 1, "")'
     )
    _set_new_value("check_need", "value", check_need_formula)

    transport_price_formula = (
        f"={ITEM_POSITION['transport_base_price'].column_letter}{row_number}*"
        f"{ITEM_POSITION['transport_numeric_ratio'].column_letter}{row_number}*"
        f"{ITEM_POSITION['gross_weight'].column_letter}{row_number}/1000"
    )
    _set_new_value("transport_price", "value", transport_price_formula)
    # 
    
    supplier_price_column = ITEM_POSITION["supplier_price"].column_letter
    transport_price_column = ITEM_POSITION["transport_price"].column_letter
    result_price_formula = (
        f"=ROUND(IF(NOT(ISBLANK({is_delivery_included_column}{row_number})), "
        f"({supplier_price_column}{row_number}-{transport_price_column}{row_number}), "
        f"{supplier_price_column}{row_number}),2)"
    )
    _set_new_value("result_price", "value", result_price_formula)
    # ic(ITEM_POSITION["result_price"].column_letter)
    # 
    last_price_column_number = ITEM_POSITION["price_history_range"].column_number + max_history_len - 1
    last_price_column = get_column_letter(last_price_column_number)
    # 
    current_selling_price_column = ITEM_POSITION["current_selling_price"].column_letter
    # 
    result_price_column = ITEM_POSITION["result_price"].column_letter
    base_price_column = ITEM_POSITION["base_price"].column_letter
    # 
    last_index_formula = f"={current_selling_price_column}{row_number}/{base_price_column}{row_number}"
    result_index_formula = f"={result_price_column}{row_number}/{base_price_column}{row_number}"
    # 
    _set_new_value("previous_index", "value", last_index_formula)
    _set_new_value("result_index", "value", result_index_formula)
    # 
    result_index_column = ITEM_POSITION["result_index"].column_letter
    previous_index_column = ITEM_POSITION["previous_index"].column_letter
    index_change_absolute_formula = (
        f'=IF({result_index_column}{row_number} - {previous_index_column}{row_number}, '
        f'{result_index_column}{row_number} - {previous_index_column}{row_number}, "")'
    )
    index_change_percentage_formula = (
        f'=IF(({result_index_column}{row_number} / {previous_index_column}{row_number}) - 1, '
        f'({result_index_column}{row_number} / {previous_index_column}{row_number}) - 1, "")'
    )

    _set_new_value("index_change_absolute", "value", index_change_absolute_formula)
    _set_new_value("index_change_in_percentage", "value", index_change_percentage_formula)
    # 
    absolute_price_change_formula = (
        f'=IF({result_price_column}{row_number}-{current_selling_price_column}{row_number}, '
        f'{result_price_column}{row_number}-{current_selling_price_column}{row_number}, "")'
    )
    percentage_price_change_formula = (
        f'=IF(({result_price_column}{row_number} / {current_selling_price_column}{row_number}) - 1, '
        f'({result_price_column}{row_number} / {current_selling_price_column}{row_number}) - 1, '
        f'"")'
    )
    # ic(percentage_price_change_formula)
    #
    _set_new_value("absolute_price_change", "value", absolute_price_change_formula)
    _set_new_value("percentage_price_change", "value", percentage_price_change_formula)
    



    d = dict(sorted(ITEM_POSITION.items(), key=lambda x: x[1].column_number))

    mod_row = [
        item
        for value in d.values()
        for item in (value.value if isinstance(value.value, list) else [value.value])
    ]

    # ic(mod_row)
    return mod_row
